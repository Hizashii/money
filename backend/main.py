"""Simple API: upload PDFs, extract invoice data, list invoices, export Excel."""
import io
from typing import List

import pdfplumber
from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from extract import InvoiceRow, extract_invoice

app = FastAPI(title="Money API", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory store (super simple)
invoices: List[dict] = []


class InvoiceOut(BaseModel):
    vendor: str
    date: str
    total: str
    vat: str


@app.post("/api/extract", response_model=List[InvoiceOut])
async def extract_pdfs(files: List[UploadFile] = File(...)):
    """Upload PDF(s), extract invoice fields, append to list and return."""
    results = []
    for f in files:
        if not f.filename or not f.filename.lower().endswith(".pdf"):
            continue
        raw = await f.read()
        text = ""
        try:
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        text += t + "\n"
        except Exception:
            text = ""
        row = extract_invoice(text, f.filename)
        obj = {"vendor": row.vendor, "date": row.date, "total": row.total, "vat": row.vat}
        invoices.append(obj)
        results.append(obj)
    return results


@app.get("/api/invoices", response_model=List[InvoiceOut])
def list_invoices():
    """Return all extracted invoices."""
    return invoices


@app.delete("/api/invoices")
def clear_invoices():
    """Clear stored invoices (optional)."""
    global invoices
    invoices = []
    return {"ok": True}


@app.get("/api/invoices/excel")
def export_excel():
    """Export all invoices to an Excel file."""
    wb = Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "Invoices"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["#", "Vendor", "Date", "Total", "VAT"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for i, inv in enumerate(invoices, 1):
        ws.cell(row=i + 1, column=1, value=i).border = border
        ws.cell(row=i + 1, column=2, value=inv["vendor"]).border = border
        ws.cell(row=i + 1, column=3, value=inv["date"]).border = border
        ws.cell(row=i + 1, column=4, value=inv["total"]).border = border
        ws.cell(row=i + 1, column=5, value=inv["vat"]).border = border

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    # Summary row
    if invoices:
        summary_row = len(invoices) + 3
        ws.cell(row=summary_row, column=3, value="Totals:").font = Font(bold=True)

        # Calculate totals
        total_sum = sum(float(inv["total"].replace(",", "") or 0) for inv in invoices if inv["total"] != "—")
        vat_sum = sum(float(inv["vat"].replace(",", "") or 0) for inv in invoices if inv["vat"] != "—")

        ws.cell(row=summary_row, column=4, value=f"{total_sum:,.2f}").font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value=f"{vat_sum:,.2f}").font = Font(bold=True)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=invoices.xlsx"},
    )
